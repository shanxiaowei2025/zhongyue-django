from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.core.paginator import Paginator
from .models import Customer
from .serializers import CustomerSerializer
from django.db.models import Q, fields as model_fields, Func, F
from django.utils import timezone
from rest_framework.pagination import PageNumberPagination
from rest_framework.viewsets import ModelViewSet
import os
from django.conf import settings
import json
from pathlib import Path
import uuid
from django.utils.encoding import escape_uri_path
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
from django.db import models
from django.db.models.functions import Lower
from apps.users.views import get_user_permissions_helper

# Create your views here.

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class CustomerViewSet(ModelViewSet):
    queryset = Customer.objects.all().order_by('-id')
    serializer_class = CustomerSerializer
    pagination_class = StandardResultsSetPagination

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

def apply_permission_filters(queryset, user):
    user_permissions = get_user_permissions_helper(user)
    customer_permissions = user_permissions['permissions']['customer']
    
    if not customer_permissions['data']['view_all']:
        filters = Q()
        
        if customer_permissions['data']['view_own']:
            filters |= Q(submitter=user.username)
        
        if customer_permissions['data']['view_by_location']:
            filters |= Q(business_address__icontains=customer_permissions['data']['view_by_location'])
        
        if customer_permissions['data']['view_department_submissions']:
            department_users = user.objects.filter(dept_id=user.dept_id).values_list('username', flat=True)
            filters |= Q(submitter__in=department_users)
        
        queryset = queryset.filter(filters)
    
    return queryset, customer_permissions

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_list(request):
    queryset = Customer.objects.all()
    queryset, customer_permissions = apply_permission_filters(queryset, request.user)


    # 创建字段映射
    field_mapping = generate_field_mapping()

    # 动态处理搜索参数
    for key, value in request.query_params.items():
        if value:
            db_field = field_mapping.get(key, key)  # 如果没有映射，使用原始键
            if hasattr(Customer, db_field):
                queryset = queryset.filter(**{f"{db_field}__icontains": value})


    # 获取查询参数
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 10))

    # 排序
    queryset = queryset.order_by('-id')

    # 分页
    paginator = Paginator(queryset, page_size)
    customers = paginator.get_page(page)

    serializer = CustomerSerializer(customers, many=True, context={'request': request})
    print(serializer.data)
    print(customer_permissions)
    return Response({
        'success': True,
        'data': {
            'list': serializer.data,
            'total': paginator.count,
            'currentPage': page,
            'pageSize': page_size,
            'permissions': customer_permissions
        }
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def create_customer(request):
    data = request.data.copy()
    serializer = CustomerSerializer(data=data)
    if serializer.is_valid():
        serializer.save(submitter=request.user.username)
        return Response({'success': True, 'data': serializer.data}, 
                      status=status.HTTP_201_CREATED)
    else:
        return Response({'success': False, 'errors': serializer.errors}, 
                      status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def update_customer(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return Response({'success': False, 'error': 'Customer not found'}, 
                      status=status.HTTP_404_NOT_FOUND)

    data = request.data.copy()
    try:
        serializer = CustomerSerializer(customer, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data})
        else:
            return Response({'success': False, 'errors': serializer.errors}, 
                          status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'success': False, 'errors': str(e)}, 
                      status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_customer(request, pk):
    try:
        customer = Customer.objects.get(pk=pk)
    except Customer.DoesNotExist:
        return Response({'success': False, 'error': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)

    customer.delete()
    return Response({'success': True}, status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_detail(request, id):
    try:
        customer = Customer.objects.get(id=id)
        serializer = CustomerSerializer(customer)
        return Response({
            "success": True,
            "data": serializer.data
        })
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "error": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_related_customers(request):
    boss_name = request.query_params.get('boss_name', None)
    
    if not boss_name:
        return Response({'success': False, 'error': 'Boss name is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    related_customers = Customer.objects.filter(boss_name=boss_name).values('id', 'company_name')
    
    if not related_customers:
        return Response({'success': False, 'error': 'No related customers found'}, status=status.HTTP_404_NOT_FOUND)
    
    return Response({
        'success': True,
        'data': list(related_customers)
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_customers(request):
    # 获取查询参数
    company_name = request.query_params.get('companyName')
    daily_contact = request.query_params.get('dailyContact')
    sales_representative = request.query_params.get('salesRepresentative')
    tax_bureau = request.query_params.get('taxBureau')
    boss_name = request.query_params.get('bossName')

    # 应用过滤
    queryset = Customer.objects.all()
    if company_name:
        queryset = queryset.filter(company_name__icontains=company_name)
    if daily_contact:
        queryset = queryset.filter(daily_contact__icontains=daily_contact)
    if sales_representative:
        queryset = queryset.filter(sales_representative__icontains=sales_representative)
    if tax_bureau:
        queryset = queryset.filter(tax_bureau__icontains=tax_bureau)
    if boss_name:
        queryset = queryset.filter(boss_name__icontains=boss_name)

    # 应用权限过滤
    queryset, _ = apply_permission_filters(queryset, request.user)

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "客户列表"

    # 获取所有字段
    fields = [field for field in Customer._meta.get_fields() if not field.is_relation]

    # 写入表头
    for col, field in enumerate(fields, start=1):
        ws.cell(row=1, column=col, value=field.verbose_name)

    # 写入数据
    for row, customer in enumerate(queryset, start=2):
        for col, field in enumerate(fields, start=1):
            value = getattr(customer, field.name)
            if isinstance(field, models.JSONField):
                value = json.dumps(value, ensure_ascii=False)
            elif isinstance(field, models.DateTimeField):
                value = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
            elif isinstance(field, models.DateField):
                value = value.strftime('%Y-%m-%d') if value else ''
            elif field.name in ['has_online_banking', 'is_online_banking_custodian']:
                value = '是' if value == 'true' else '否' if value == 'false' else ''
            ws.cell(row=row, column=col, value=value)

    # 调整列宽
    for col in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 创建 HTTP 响应
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

    return response

def generate_field_mapping():
    """
    生成前端驼峰命名到后端下划线命名的字段映射
    """
    def snake_to_camel(name):
        components = name.split('_')
        return components[0] + ''.join(x.title() for x in components[1:])

    # 获取模型的所有字段
    model_field_names = [f.name for f in Customer._meta.get_fields() if isinstance(f, model_fields.Field)]
    
    # 创建映射
    return {snake_to_camel(field): field for field in model_field_names}

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_autocomplete_options(request):
    field = request.query_params.get('field')
    query = request.query_params.get('query', '').lower()
    
    if not field:
        return Response({'success': False, 'error': 'Field parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    if not hasattr(Customer, field):
        return Response({'success': False, 'error': 'Invalid field'}, status=status.HTTP_400_BAD_REQUEST)
    
    # 应用权限过滤
    queryset, _ = apply_permission_filters(Customer.objects.all(), request.user)
    
    # 获取唯一值，忽略大小写，并按字段值排序
    values = queryset.annotate(
        lower_field=Lower(field)
    ).filter(
        lower_field__contains=query
    ).values_list(field, flat=True).distinct().order_by(field)[:10]  # 限制返回前10个结果
    
    return Response({
        'success': True,
        'data': list(values)
    })
